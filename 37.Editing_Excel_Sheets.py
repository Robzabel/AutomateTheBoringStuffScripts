import openpyxl, os
wb = openpyxl.Workbook() #make sure to have a capital W, this creates the workbook
#previously we have uses wb.get_sheet_names to return the names of the worksheets but now we should use:
print(wb.sheetnames)#print the names of the worksheets in the WB
#previously we have used wb.get_sheet_by_name but now we should use:
sheet = wb['Sheet'] # create a variable to access the sheet
sheet['A1'] = 42 # we can now use the sheet variable to access the cells
sheet['A2'] = 'Hello World'
os.chdir('/home/zabex/Desktop/')
sheet2 = wb.create_sheet('Sheet2') #create another sheet in the wb and save it to a variable
print(wb.sheetnames)
sheet.title = 'Sheet1'# Change the name of a sheet
print(wb.sheetnames)
sheet0 = wb.create_sheet(index=0, title='Sheet0')#create a new sheet and put it at index 0
print(wb.sheetnames)
wb.save('FirstExample.xlsx')
sheet.